from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from math import isnan
from re import sub
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from xlrd import open_workbook

CANONICAL_ALIASES = {
    "sku": "sku",
    "title": "title",
    "name": "title",
    "bullet points": "bullet_points",
    "bullet_points": "bullet_points",
    "specs": "specs",
    "category": "category",
    "product type": "product_type",
    "product_type": "product_type",
    "attributes lulu product type": "product_type",
    "attribute set": "attribute_set",
    "attribute_set": "attribute_set",
    "search query": "search_query",
    "search_query": "search_query",
    "l1": "l1",
    "l2": "l2",
    "l3": "l3",
    "l4": "l4",
}
CORE_FIELDS = set(CANONICAL_ALIASES.values())


@dataclass
class ParsedRow:
    row_number: int
    sku: str
    core: dict[str, Any]
    attributes: dict[str, Any]
    source_row: dict[str, Any]


@dataclass
class ParseError:
    row_number: int
    sku: str | None
    error_code: str
    message: str
    field_header: str | None = None


@dataclass
class ParseResult:
    total_rows: int
    valid_rows: list[ParsedRow]
    errors: list[ParseError]
    duplicate_skus: list[dict[str, Any]]


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = sub(r"\s+", " ", text.strip().replace("_", " "))
    return text.casefold()


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and isnan(value):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def normalize_sku(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_rows(raw_headers: tuple[Any, ...] | list[Any], rows: Any, max_rows: int) -> ParseResult:
    headers = ["" if header is None else str(header).strip() for header in raw_headers]
    canonical_by_index = [CANONICAL_ALIASES.get(normalize_header(header)) for header in headers]
    if "sku" not in canonical_by_index:
        raise ValueError("Missing SKU header")

    sku_index = canonical_by_index.index("sku")
    parsed: list[ParsedRow] = []
    errors: list[ParseError] = []
    seen: dict[str, list[int]] = defaultdict(list)
    total_rows = 0

    for excel_row_number, row_values in rows:
        if total_rows >= max_rows:
            raise ValueError(f"Workbook exceeds configured row limit of {max_rows}")
        values = list(row_values)
        if not any(normalize_cell(value) is not None for value in values):
            continue
        total_rows += 1
        source_row: dict[str, Any] = {}
        attributes: dict[str, Any] = {}
        core: dict[str, Any] = {}

        for index, header in enumerate(headers):
            if not header:
                continue
            value = normalize_cell(values[index] if index < len(values) else None)
            source_row[header] = value
            canonical = canonical_by_index[index]
            if canonical:
                core[canonical] = normalize_sku(value) if canonical == "sku" else value
            else:
                attributes[header] = value

        sku = normalize_sku(values[sku_index] if sku_index < len(values) else None)
        if not sku:
            errors.append(ParseError(excel_row_number, None, "missing_sku", "SKU is required", headers[sku_index]))
            continue
        seen[sku].append(excel_row_number)
        parsed.append(ParsedRow(excel_row_number, sku, core, attributes, source_row))

    duplicate_skus = [{"sku": sku, "rows": rows} for sku, rows in seen.items() if len(rows) > 1]
    duplicate_set = {item["sku"] for item in duplicate_skus}
    if duplicate_set:
        for item in duplicate_skus:
            for row_number in item["rows"]:
                errors.append(
                    ParseError(row_number, item["sku"], "duplicate_sku", "Duplicate SKU in workbook", headers[sku_index])
                )
        parsed = [row for row in parsed if row.sku not in duplicate_set]

    return ParseResult(total_rows, parsed, errors, duplicate_skus)


def parse_xlsx(content: bytes, filename: str, max_rows: int) -> ParseResult:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except (BadZipFile, KeyError, OSError) as exc:
            raise ValueError("Invalid .xlsx workbook") from exc

        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError("Workbook is empty") from exc
        return _parse_rows(raw_headers, enumerate(rows, start=2), max_rows)

    if lower.endswith(".xls"):
        try:
            workbook = open_workbook(file_contents=content, on_demand=True)
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            raise ValueError("Invalid .xls workbook") from exc
        if sheet.nrows == 0:
            raise ValueError("Workbook is empty")
        rows = ((index + 1, sheet.row_values(index)) for index in range(1, sheet.nrows))
        return _parse_rows(sheet.row_values(0), rows, max_rows)

    raise ValueError("Only .xlsx or .xls files are accepted")
